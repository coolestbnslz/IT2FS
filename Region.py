import xlrd
import openpyxl
import numpy as np
import matplotlib.pyplot as plt
file_name ="C:\\Users\\Nikhil Bansal\\Desktop\\linguistic_ratings_new.xlsx"
import pandas as pd
wb1 = openpyxl.load_workbook(filename=file_name)
ws = wb1.worksheets[0]
wb = xlrd.open_workbook(file_name)
sheet = wb.sheet_by_index(0)
df=pd.read_excel((file_name),sheet_name='data',index_col=None)
def line(a,b,c,p):
    import matplotlib.pyplot as plt
    if c==0:
        y1=np.zeros(1001)
        y2=np.zeros(1001)
        d=max(a.dropna())
        e=max(b.dropna())
        x_final=np.arange(0,10.01,0.01)
        for i in range(len(x_final)):
            if x_final[i]<=e:
                y1[i] = (-x_final[i] / e) + 1
            else:
                y1[i]=0
        for i in range(len(x_final)):
            if x_final[i]<=d:
                y2[i] = (-x_final[i] / d) + 1
            else:
                y2[i]=0
        plt.plot(x_final,y1)
        plt.plot(x_final,y2)
        plt.xlim(xmin=0, xmax=10)
        plt.ylim(ymin=0, ymax=1)
        plt.xticks(np.arange(0,10.5,0.5))
        plt.xlabel(str(ws.cell(row=1,column=p).value))
        plt.show()
        return y1,y2
    if c==2:
        y1 = np.zeros(1001)
        y2 = np.zeros(1001)
        d = min(a.dropna())
        e = min(b.dropna())
        x_final = np.arange(0, 10.01, 0.01)
        for i in range(len(x_final)):
            if x_final[i] >= e:
                y2[i] = (x_final[i]-e)/(10-e)
            else:
                y2[i] = 0
        for i in range(len(x_final)):
            if x_final[i] >= d:
                y1[i] = (x_final[i]-d)/(10-d)
            else:
                y1[i] = 0
        plt.plot(x_final, y1)
        plt.plot(x_final, y2)
        plt.xticks(np.arange(0, 10.5, 0.5))
        plt.xlim(xmin=0, xmax=10)
        plt.ylim(ymin=0, ymax=1)
        plt.xlabel(str(ws.cell(row=1, column=p).value))
        plt.show()
        return y1,y2
    if c==1:
        y1 = np.zeros(1001)
        y2 = np.zeros(1001)
        x_final = np.arange(0, 10.01, 0.01)
        d = min(a.dropna())
        e = max(b.dropna())
        g=max(a.dropna())
        h=min(b.dropna())
        f=(d+e)/2
        for i in range(len(x_final)):
            if x_final[i] >= d and x_final[i]<=f:
                y1[i] = (x_final[i]-d)/(f-d)
            elif x_final[i]>f and x_final[i]<=e:
                y1[i] = (x_final[i] - e) / (f - e)
        for i in range(len(x_final)):
            if x_final[i] >= g and x_final[i] <= f:
                y2[i] = (x_final[i] - g) / (f - g)
            elif x_final[i] > f and x_final[i] <= h:
                y2[i] = (x_final[i] - h) / (f - h)
        plt.plot(x_final, y1)
        plt.plot(x_final, y2)
        plt.xlim(xmin=0, xmax=10)
        plt.ylim(ymin=0, ymax=1)
        plt.xticks(np.arange(0, 10.5, 0.5))
        plt.xlabel(str(ws.cell(row=1, column=p).value))
        plt.show()
        return y1,y2



label_dict={0:'left', 1: 'interior', 2: 'right'}
extra=[0,0,0]
label=[]
a=df['VL']
b=df['Unnamed: 1']
c=5.831*a-b
d=(0.171*a)+8.29-b
for i in range(len(a)):
    if d[i]>=0 and c[i]>=0:
        label.append(1)
        extra[1]=extra[1]+1
    elif c[i]<0 and d[i]>=0:
        label.append(0)
        extra[0]=extra[0]+1
    elif c[i]>=0 and d[i]<0:
        label.append(2)
        extra[2]=extra[2]+1
    else:
        label.append(3)
e=np.argmax(extra)
print(label_dict[e])
VL_upper,VL_lower=line(a,b,e,1)

for i in range(len(a)):
    if label[i]!=e:
        ws.cell(row=i + 2, column=1).value = None
        ws.cell(row=i + 2, column=2).value = None
extra=[0,0,0]
label=[]
a=df['L']
b=df['Unnamed: 3']
c=5.831*a-b
d=(0.171*a)+8.29-b
for i in range(len(a)):
    if d[i]>=0 and c[i]>=0:
        label.append(1)
        extra[1]=extra[1]+1
    elif c[i]<0 and d[i]>=0:
        label.append(0)
        extra[0]=extra[0]+1
    elif c[i]>=0 and d[i]<0:
        label.append(2)
        extra[2]=extra[2]+1
    else:
        label.append(3)
e=np.argmax(extra)
print(label_dict[e])
L_upper,L_lower=line(a,b,e,3)
for i in range(len(a)):
    if label[i]!=e:
        ws.cell(row=i + 2, column=3).value = None
        ws.cell(row=i + 2, column=4).value = None
extra=[0,0,0]
label=[]
a=df['M']
b=df['Unnamed: 5']
c=5.831*a-b
d=(0.171*a)+8.29-b
for i in range(len(a)):
    if d[i]>=0 and c[i]>=0:
        label.append(1)
        extra[1]=extra[1]+1
    elif c[i]<0 and d[i]>=0:
        label.append(0)
        extra[0]=extra[0]+1
    elif c[i]>=0 and d[i]<0:
        label.append(2)
        extra[2]=extra[2]+1
    else:
        label.append(3)
e=np.argmax(extra)
print(label_dict[e])
M_upper,M_lower=line(a,b,e,5)
for i in range(len(a)):
    if label[i]!=e:
        ws.cell(row=i + 2, column=5).value = None
        ws.cell(row=i + 2, column=6).value = None
extra=[0,0,0]
label=[]
a=df['LH']
b=df['Unnamed: 7']
c=5.831*a-b
d=(0.171*a)+8.29-b
for i in range(len(a)):
    if d[i]>=0 and c[i]>=0:
        label.append(1)
        extra[1]=extra[1]+1
    elif c[i]<0 and d[i]>=0:
        label.append(0)
        extra[0]=extra[0]+1
    elif c[i]>=0 and d[i]<0:
        label.append(2)
        extra[2]=extra[2]+1
    else:
        label.append(3)
e=np.argmax(extra)
print(label_dict[e])
LH_upper,LH_lower=line(a,b,e,7)
for i in range(len(a)):
    if label[i]!=e:
        ws.cell(row=i + 2, column=7).value = None
        ws.cell(row=i + 2, column=8).value = None
extra=[0,0,0]
label=[]
a=df['H']
b=df['Unnamed: 9']
c=5.831*a-b
d=(0.171*a)+8.29-b
for i in range(len(a)):
    if d[i]>=0 and c[i]>=0:
        label.append(1)
        extra[1]=extra[1]+1
    elif c[i]<0 and d[i]>=0:
        label.append(0)
        extra[0]=extra[0]+1
    elif c[i]>=0 and d[i]<0:
        label.append(2)
        extra[2]=extra[2]+1
    else:
        label.append(3)
e=np.argmax(extra)
print(label_dict[e])
H_upper,H_lower=line(a,b,e,9)
for i in range(len(a)):
    if label[i]!=e:
        ws.cell(row=i + 2, column=9).value = None
        ws.cell(row=i + 2, column=10).value = None
extra=[0,0,0]
label=[]
a=df['VH']
b=df['Unnamed: 11']
c=5.831*a-b
d=(0.171*a)+8.29-b
for i in range(len(a)):
    if d[i]>=0 and c[i]>=0:
        label.append(1)
        extra[1]=extra[1]+1
    elif c[i]<0 and d[i]>=0:
        label.append(0)
        extra[0]=extra[0]+1
    elif c[i]>=0 and d[i]<0:
        label.append(2)
        extra[2]=extra[2]+1
    else:
        label.append(3)
e=np.argmax(extra)
print(label_dict[e])
VH_upper,VH_lower=line(a,b,e,11)
for i in range(len(a)):
    if label[i]!=e:
        ws.cell(row=i + 2, column=11).value = None
        ws.cell(row=i + 2, column=12).value = None
extra=[0,0,0]
label=[]
a=df['P']
b=df['Unnamed: 13']
c=5.831*a-b
d=(0.171*a)+8.29-b
for i in range(len(a)):
    if d[i]>=0 and c[i]>=0:
        label.append(1)
        extra[1]=extra[1]+1
    elif c[i]<0 and d[i]>=0:
        label.append(0)
        extra[0]=extra[0]+1
    elif c[i]>=0 and d[i]<0:
        label.append(2)
        extra[2]=extra[2]+1
    else:
        label.append(3)
e=np.argmax(extra)
print(label_dict[e])
P_upper,P_lower=line(a,b,e,13)
for i in range(len(a)):
    if label[i]!=e:
        ws.cell(row=i + 2, column=13).value = None
        ws.cell(row=i + 2, column=14).value = None
extra=[0,0,0]
label=[]
a=df['MLP']
b=df['Unnamed: 15']
c=5.831*a-b
d=(0.171*a)+8.29-b
for i in range(len(a)):
    if d[i]>=0 and c[i]>=0:
        label.append(1)
        extra[1]=extra[1]+1
    elif c[i]<0 and d[i]>=0:
        label.append(0)
        extra[0]=extra[0]+1
    elif c[i]>=0 and d[i]<0:
        label.append(2)
        extra[2]=extra[2]+1
    else:
        label.append(3)
e=np.argmax(extra)
print(label_dict[e])
MLP_upper,MLP_lower=line(a,b,e,15)
for i in range(len(a)):
    if label[i]!=e:
        ws.cell(row=i + 2, column=15).value = None
        ws.cell(row=i + 2, column=16).value = None
extra=[0,0,0]
label=[]
a=df['ME']
b=df['Unnamed: 17']
c=5.831*a-b
d=(0.171*a)+8.29-b
for i in range(len(a)):
    if d[i]>=0 and c[i]>=0:
        label.append(1)
        extra[1]=extra[1]+1
    elif c[i]<0 and d[i]>=0:
        label.append(0)
        extra[0]=extra[0]+1
    elif c[i]>=0 and d[i]<0:
        label.append(2)
        extra[2]=extra[2]+1
    else:
        label.append(3)
e=np.argmax(extra)
print(label_dict[e])
ME_upper,ME_lower=line(a,b,e,17)
for i in range(len(a)):
    if label[i]!=e:
        ws.cell(row=i + 2, column=17).value = None
        ws.cell(row=i + 2, column=18).value = None
extra=[0,0,0]
label=[]
a=df['G']
b=df['Unnamed: 19']
c=5.831*a-b
d=(0.171*a)+8.29-b
for i in range(len(a)):
    if d[i]>=0 and c[i]>=0:
        label.append(1)
        extra[1]=extra[1]+1
    elif c[i]<0 and d[i]>=0:
        label.append(0)
        extra[0]=extra[0]+1
    elif c[i]>=0 and d[i]<0:
        label.append(2)
        extra[2]=extra[2]+1
    else:
        label.append(3)
e=np.argmax(extra)
print(label_dict[e])
G_upper,G_lower=line(a,b,e,19)
for i in range(len(a)):
    if label[i]!=e:
        ws.cell(row=i + 2, column=19).value = None
        ws.cell(row=i + 2, column=20).value = None
extra=[0,0,0]
label=[]
a=df['VG']
b=df['Unnamed: 21']
c=5.831*a-b
d=(0.171*a)+8.29-b
for i in range(len(a)):
    if d[i]>=0 and c[i]>=0:
        label.append(1)
        extra[1]=extra[1]+1
    elif c[i]<0 and d[i]>=0:
        label.append(0)
        extra[0]=extra[0]+1
    elif c[i]>=0 and d[i]<0:
        label.append(2)
        extra[2]=extra[2]+1
    else:
        label.append(3)
e=np.argmax(extra)
print(label_dict[e])
VG_upper,VG_lower=line(a,b,e,21)
for i in range(len(a)):
    if label[i]!=e:
        ws.cell(row=i + 2, column=21).value = None
        ws.cell(row=i + 2, column=22).value = None
wb1.save(file_name)
"""
considering VL,L,M,VG,VH only
"""
x_final=(np.arange(0,10.01,0.01))
plt.fill_between(x_final,VL_upper,VL_lower,label='VL')
plt.fill_between(x_final,L_upper,L_lower,label='L')
plt.fill_between(x_final,M_upper,M_lower,label='M')
plt.fill_between(x_final,VG_upper,VG_lower,label='VG')
plt.fill_between(x_final,VH_upper,VH_lower,label='VH')
plt.xlim(xmin=0, xmax=10)
plt.ylim(ymin=0, ymax=1)
plt.xticks(np.arange(0, 10.5, 0.5))
plt.xlabel('Quality')
plt.legend()
plt.show()

quality=float(input('enter the value of quality='))
price=float(input('enter the value of price='))
def matrix(VL_lower,L_lower,M_lower,VG_lower,VH_lower,quality):
    lower=[]
    c=quality*100
    lower.append(VL_lower[int(c+1)])
    lower.append(L_lower[int(c+ 1)])
    lower.append(M_lower[int(c+ 1)])
    lower.append(VG_lower[int(c+ 1)])
    lower.append(VH_lower[int(c+ 1)])
    return lower
lower_quality=matrix(VL_lower,L_lower,M_lower,VG_lower,VH_lower,quality)
upper_quality=matrix(VL_upper,L_upper,M_upper,VG_upper,VH_upper,quality)
lower_price=matrix(VL_lower,L_lower,M_lower,VG_lower,VH_lower,price)
upper_price=matrix(VL_upper,L_upper,M_upper,VG_upper,VH_upper,price)
lower_matrix=np.zeros((5,5))
upper_matrix=np.zeros((5,5))
for i in range(5):
    for j in range(5):
        lower_matrix[i][j]=min(lower_quality[i],lower_price[j])
for i in range(5):
    for j in range(5):
        upper_matrix[i][j]=min(upper_quality[i],upper_price[j])
rules=[[1,1,1,1,1],[2,2,2,1,1],[3,3,3,3,2],[4,4,4,4,3],[5,5,5,5,5]]
dict_lower={}
for i in range(len(lower_matrix)):
    for j in range(len(lower_matrix[0])):
        dict_lower.setdefault(rules[i][j],[]).append(lower_matrix[i][j])
dict_upper={}
for i in range(len(lower_matrix)):
    for j in range(len(lower_matrix[0])):
        dict_upper.setdefault(rules[i][j],[]).append(upper_matrix[i][j])
print(dict_lower)
print(dict_upper)
vl_lower_value=max(dict_lower[1])
l_lower_value=max(dict_lower[2])
m_lower_value=max(dict_lower[3])
vg_lower_value=max(dict_lower[4])
vh_lower_value=max(dict_lower[5])
vl_upper_value=max(dict_upper[1])
l_upper_value=max(dict_upper[2])
m_upper_value=max(dict_upper[3])
vg_upper_value=max(dict_upper[4])
vh_upper_value=max(dict_upper[5])
def clipping(lower,upper,lower_value,upper_value):
    for i in range(len(lower)):
            if lower[i]>=lower_value:
                lower[i]=lower_value
    for i in range(len(upper)):
            if upper[i]>=upper_value:
                upper[i]=upper_value
clipping(VL_lower,VL_upper,vl_lower_value,vl_upper_value)
clipping(L_lower,L_upper,l_lower_value,l_upper_value)
clipping(M_lower,M_upper,m_lower_value,m_upper_value)
clipping(VG_lower,VG_upper,vg_lower_value,vg_upper_value)
clipping(VH_lower,VH_upper,vh_lower_value,vh_upper_value)
plt.fill_between(x_final,VL_upper,VL_lower,label='VL')
plt.fill_between(x_final,L_upper,L_lower,label='L')
plt.fill_between(x_final,M_upper,M_lower,label='M')
plt.fill_between(x_final,VG_upper,VG_lower,label='VG')
plt.fill_between(x_final,VH_upper,VH_lower,label='VH')
plt.xlim(xmin=0, xmax=10)
plt.ylim(ymin=0, ymax=1)
plt.xticks(np.arange(0, 10.5, 0.5))
plt.xlabel('Rating')
plt.legend()
plt.show()
y_final_upper=np.zeros(1001)
y_final_lower=np.zeros(1001)
for i in range(len(y_final_upper)):
    y_final_upper[i]=max(VL_upper[i],L_upper[i],M_upper[i],VG_upper[i],VH_upper[i])
    y_final_lower[i]=max(VL_lower[i],L_lower[i],M_lower[i],VG_lower[i],VH_lower[i])

plt.fill_between(x_final,y_final_upper,y_final_lower)
plt.xlim(xmin=0, xmax=10)
plt.ylim(ymin=0, ymax=1)
plt.xticks(np.arange(0, 10.5, 0.5))
plt.xlabel('Rating')
plt.show()
rating_value_upper=sum(x_final*y_final_upper)/sum(y_final_upper)
rating_value_lower=sum(x_final*y_final_lower)/sum(y_final_lower)
print(rating_value_lower)
print(rating_value_upper)