import xlrd
import openpyxl
import numpy as np
import statistics
from numpy import log as ln
from decimal import Decimal, getcontext
from fractions import Fraction

getcontext().prec = 2*10
file_name ="C:\\Users\\Nikhil Bansal\\Desktop\\linguistic_ratings_new.xlsx"

import pandas as pd
wb1 = openpyxl.load_workbook(filename=file_name)
ws = wb1.worksheets[0]
wb = xlrd.open_workbook(file_name)
sheet = wb.sheet_by_index(0)
df=pd.read_excel((file_name),sheet_name='data',index_col=None)
print(df)
a=df['VL']
b=df['Unnamed: 1']

for i in range(len(a)):
    if a[i]>b[i] or a[i]>10 or b[i]>10:
        ws.cell(row=i+2,column=1).value=None
        ws.cell(row=i+2,column=2).value =None
a=df['L']
b=df['Unnamed: 3']
for i in range(len(a)):
    if a[i]>b[i] or a[i]>10 or b[i]>10:
        ws.cell(row=i+2,column=3).value=None
        ws.cell(row=i+2,column=4).value =None
a=df['M']
b=df['Unnamed: 5']
for i in range(len(a)):
    if a[i]>b[i] or a[i]>10 or b[i]>10:
        ws.cell(row=i+2,column=5).value=None
        ws.cell(row=i+2,column=6).value =None
a=df['LH']
b=df['Unnamed: 7']
for i in range(len(a)):
    if a[i]>b[i] or a[i]>10 or b[i]>10:
        ws.cell(row=i+2,column=7).value=None
        ws.cell(row=i+2,column=8).value =None
a=df['H']
b=df['Unnamed: 9']
for i in range(len(a)):
    if a[i]>b[i] or a[i]>10 or b[i]>10:
        ws.cell(row=i+2,column=9).value=None
        ws.cell(row=i+2,column=10).value =None

a=df['VH']
b=df['Unnamed: 11']
for i in range(len(a)):
    if a[i]>b[i] or a[i]>10 or b[i]>10:
        ws.cell(row=i+2,column=11).value=None
        ws.cell(row=i+2,column=12).value =None
a=df['P']
b=df['Unnamed: 13']
for i in range(len(a)):
    if a[i]>b[i] or a[i]>10 or b[i]>10:
        ws.cell(row=i+2,column=13).value=None
        ws.cell(row=i+2,column=14).value =None
a=df['MLP']
b=df['Unnamed: 15']
for i in range(len(a)):
    if a[i]>b[i] or a[i]>10 or b[i]>10:
        ws.cell(row=i+2,column=15).value=None
        ws.cell(row=i+2,column=16).value =None
a=df['ME']
b=df['Unnamed: 17']
for i in range(len(a)):
    if a[i]>b[i] or a[i]>10 or b[i]>10:
        ws.cell(row=i+2,column=17).value=None
        ws.cell(row=i+2,column=18).value =None
a=df['G']
b=df['Unnamed: 19']
for i in range(len(a)):
    if a[i]>b[i] or a[i]>10 or b[i]>10:
        ws.cell(row=i+2,column=19).value=None
        ws.cell(row=i+2,column=20).value =None
a=df['VG']
b=df['Unnamed: 21']
for i in range(len(a)):
    if a[i]>b[i] or a[i]>10 or b[i]>10:
        ws.cell(row=i+2,column=21).value=None
        ws.cell(row=i+2,column=22).value =None
wb1.save(file_name)
a=df['VL']
b=df['Unnamed: 1']
l=b-a
a_quartile=np.percentile(a.dropna(), [25, 50, 75])
b_quartile=np.percentile(b.dropna(), [25, 50, 75])
l_quartile=np.percentile(l.dropna(),[25,50,75])
for i in range(len(a)):
    if a[i]<a_quartile[0]-1.5*(a_quartile[2]-a_quartile[0]) or a[i]>a_quartile[0]+1.5*(a_quartile[2]-a_quartile[0]) or b[i]<b_quartile[0]-1.5*(b_quartile[2]-b_quartile[0]) or b[i]>b_quartile[0]+1.5*(b_quartile[2]-b_quartile[0]):
        ws.cell(row=i+2,column=1).value=None
        ws.cell(row=i+2,column=2).value =None
    elif l[i]<l_quartile[0]-1.5*(l_quartile[2]-l_quartile[0]) or l[i]>l_quartile[0]+1.5*(l_quartile[2]-l_quartile[0]):
        ws.cell(row=i + 2, column=1).value = None
        ws.cell(row=i + 2, column=2).value = None
a=df['L']
b=df['Unnamed: 3']
l=b-a
a_quartile=np.percentile(a.dropna(), [25, 50, 75])
b_quartile=np.percentile(b.dropna(), [25, 50, 75])
l_quartile=np.percentile(l.dropna(),[25,50,75])
for i in range(len(a)):
    if a[i]<a_quartile[0]-1.5*(a_quartile[2]-a_quartile[0]) or a[i]>a_quartile[0]+1.5*(a_quartile[2]-a_quartile[0]) or b[i]<b_quartile[0]-1.5*(b_quartile[2]-b_quartile[0]) or b[i]>b_quartile[0]+1.5*(b_quartile[2]-b_quartile[0]):
        ws.cell(row=i+2,column=3).value=None
        ws.cell(row=i+2,column=4).value =None
    elif l[i]<l_quartile[0]-1.5*(l_quartile[2]-l_quartile[0]) or l[i]>l_quartile[0]+1.5*(l_quartile[2]-l_quartile[0]):
        ws.cell(row=i + 2, column=3).value = None
        ws.cell(row=i + 2, column=4).value = None
a=df['M']
b=df['Unnamed: 5']
l=b-a
a_quartile=np.percentile(a.dropna(), [25, 50, 75])
b_quartile=np.percentile(b.dropna(), [25, 50, 75])
l_quartile=np.percentile(l.dropna(),[25,50,75])
for i in range(len(a)):
    if a[i]<a_quartile[0]-1.5*(a_quartile[2]-a_quartile[0]) or a[i]>a_quartile[0]+1.5*(a_quartile[2]-a_quartile[0]) or b[i]<b_quartile[0]-1.5*(b_quartile[2]-b_quartile[0]) or b[i]>b_quartile[0]+1.5*(b_quartile[2]-b_quartile[0]):
        ws.cell(row=i+2,column=5).value=None
        ws.cell(row=i+2,column=6).value =None
    elif l[i]<l_quartile[0]-1.5*(l_quartile[2]-l_quartile[0]) or l[i]>l_quartile[0]+1.5*(l_quartile[2]-l_quartile[0]):
        ws.cell(row=i + 2, column=5).value = None
        ws.cell(row=i + 2, column=6).value = None
a=df['LH']
b=df['Unnamed: 7']
l=b-a
a_quartile=np.percentile(a.dropna(), [25, 50, 75])
b_quartile=np.percentile(b.dropna(), [25, 50, 75])
l_quartile=np.percentile(l.dropna(),[25,50,75])
for i in range(len(a)):
    if a[i]<a_quartile[0]-1.5*(a_quartile[2]-a_quartile[0]) or a[i]>a_quartile[0]+1.5*(a_quartile[2]-a_quartile[0]) or b[i]<b_quartile[0]-1.5*(b_quartile[2]-b_quartile[0]) or b[i]>b_quartile[0]+1.5*(b_quartile[2]-b_quartile[0]):
        ws.cell(row=i+2,column=7).value=None
        ws.cell(row=i+2,column=8).value =None
    elif l[i]<l_quartile[0]-1.5*(l_quartile[2]-l_quartile[0]) or l[i]>l_quartile[0]+1.5*(l_quartile[2]-l_quartile[0]):
        ws.cell(row=i + 2, column=7).value = None
        ws.cell(row=i + 2, column=8).value = None
a=df['H']
b=df['Unnamed: 9']
l=b-a
a_quartile=np.percentile(a.dropna(), [25, 50, 75])
b_quartile=np.percentile(b.dropna(), [25, 50, 75])
l_quartile=np.percentile(l.dropna(),[25,50,75])
for i in range(len(a)):
    if a[i]<a_quartile[0]-1.5*(a_quartile[2]-a_quartile[0]) or a[i]>a_quartile[0]+1.5*(a_quartile[2]-a_quartile[0]) or b[i]<b_quartile[0]-1.5*(b_quartile[2]-b_quartile[0]) or b[i]>b_quartile[0]+1.5*(b_quartile[2]-b_quartile[0]):
        ws.cell(row=i+2,column=9).value=None
        ws.cell(row=i+2,column=10).value =None
    elif l[i]<l_quartile[0]-1.5*(l_quartile[2]-l_quartile[0]) or l[i]>l_quartile[0]+1.5*(l_quartile[2]-l_quartile[0]):
        ws.cell(row=i + 2, column=9).value = None
        ws.cell(row=i + 2, column=10).value = None
a=df['VH']
b=df['Unnamed: 11']
l=b-a
a_quartile=np.percentile(a.dropna(), [25, 50, 75])
b_quartile=np.percentile(b.dropna(), [25, 50, 75])
l_quartile=np.percentile(l.dropna(),[25,50,75])
for i in range(len(a)):
    if a[i]<a_quartile[0]-1.5*(a_quartile[2]-a_quartile[0]) or a[i]>a_quartile[0]+1.5*(a_quartile[2]-a_quartile[0]) or b[i]<b_quartile[0]-1.5*(b_quartile[2]-b_quartile[0]) or b[i]>b_quartile[0]+1.5*(b_quartile[2]-b_quartile[0]):
        ws.cell(row=i+2,column=11).value=None
        ws.cell(row=i+2,column=12).value =None
    elif l[i]<l_quartile[0]-1.5*(l_quartile[2]-l_quartile[0]) or l[i]>l_quartile[0]+1.5*(l_quartile[2]-l_quartile[0]):
        ws.cell(row=i + 2, column=11).value = None
        ws.cell(row=i + 2, column=12).value = None
a=df['P']
b=df['Unnamed: 13']
l=b-a
a_quartile=np.percentile(a.dropna(), [25, 50, 75])
b_quartile=np.percentile(b.dropna(), [25, 50, 75])
l_quartile=np.percentile(l.dropna(),[25,50,75])
for i in range(len(a)):
    if a[i]<a_quartile[0]-1.5*(a_quartile[2]-a_quartile[0]) or a[i]>a_quartile[0]+1.5*(a_quartile[2]-a_quartile[0]) or b[i]<b_quartile[0]-1.5*(b_quartile[2]-b_quartile[0]) or b[i]>b_quartile[0]+1.5*(b_quartile[2]-b_quartile[0]):
        ws.cell(row=i+2,column=13).value=None
        ws.cell(row=i+2,column=14).value =None
    elif l[i]<l_quartile[0]-1.5*(l_quartile[2]-l_quartile[0]) or l[i]>l_quartile[0]+1.5*(l_quartile[2]-l_quartile[0]):
        ws.cell(row=i + 2, column=13).value = None
        ws.cell(row=i + 2, column=14).value = None
a=df['MLP']
b=df['Unnamed: 15']
l=b-a
a_quartile=np.percentile(a.dropna(), [25, 50, 75])
b_quartile=np.percentile(b.dropna(), [25, 50, 75])
l_quartile=np.percentile(l.dropna(),[25,50,75])
for i in range(len(a)):
    if a[i]<a_quartile[0]-1.5*(a_quartile[2]-a_quartile[0]) or a[i]>a_quartile[0]+1.5*(a_quartile[2]-a_quartile[0]) or b[i]<b_quartile[0]-1.5*(b_quartile[2]-b_quartile[0]) or b[i]>b_quartile[0]+1.5*(b_quartile[2]-b_quartile[0]):
        ws.cell(row=i+2,column=15).value=None
        ws.cell(row=i+2,column=16).value =None
    elif l[i]<l_quartile[0]-1.5*(l_quartile[2]-l_quartile[0]) or l[i]>l_quartile[0]+1.5*(l_quartile[2]-l_quartile[0]):
        ws.cell(row=i + 2, column=15).value = None
        ws.cell(row=i + 2, column=16).value = None
a=df['ME']
b=df['Unnamed: 17']
l=b-a
a_quartile=np.percentile(a.dropna(), [25, 50, 75])
b_quartile=np.percentile(b.dropna(), [25, 50, 75])
l_quartile=np.percentile(l.dropna(),[25,50,75])
for i in range(len(a)):
    if a[i]<a_quartile[0]-1.5*(a_quartile[2]-a_quartile[0]) or a[i]>a_quartile[0]+1.5*(a_quartile[2]-a_quartile[0]) or b[i]<b_quartile[0]-1.5*(b_quartile[2]-b_quartile[0]) or b[i]>b_quartile[0]+1.5*(b_quartile[2]-b_quartile[0]):
        ws.cell(row=i+2,column=17).value=None
        ws.cell(row=i+2,column=18).value =None
    elif l[i]<l_quartile[0]-1.5*(l_quartile[2]-l_quartile[0]) or l[i]>l_quartile[0]+1.5*(l_quartile[2]-l_quartile[0]):
        ws.cell(row=i + 2, column=17).value = None
        ws.cell(row=i + 2, column=18).value = None
a=df['G']
b=df['Unnamed: 19']
l=b-a
a_quartile=np.percentile(a.dropna(), [25, 50, 75])
b_quartile=np.percentile(b.dropna(), [25, 50, 75])
l_quartile=np.percentile(l.dropna(),[25,50,75])
for i in range(len(a)):
    if a[i]<a_quartile[0]-1.5*(a_quartile[2]-a_quartile[0]) or a[i]>a_quartile[0]+1.5*(a_quartile[2]-a_quartile[0]) or b[i]<b_quartile[0]-1.5*(b_quartile[2]-b_quartile[0]) or b[i]>b_quartile[0]+1.5*(b_quartile[2]-b_quartile[0]):
        ws.cell(row=i+2,column=19).value=None
        ws.cell(row=i+2,column=20).value =None
    elif l[i]<l_quartile[0]-1.5*(l_quartile[2]-l_quartile[0]) or l[i]>l_quartile[0]+1.5*(l_quartile[2]-l_quartile[0]):
        ws.cell(row=i + 2, column=19).value = None
        ws.cell(row=i + 2, column=20).value = None
a=df['VG']
b=df['Unnamed: 21']
l=b-a
a_quartile=np.percentile(a.dropna(), [25, 50, 75])
b_quartile=np.percentile(b.dropna(), [25, 50, 75])
l_quartile=np.percentile(l.dropna(),[25,50,75])
for i in range(len(a)):
    if a[i]<a_quartile[0]-1.5*(a_quartile[2]-a_quartile[0]) or a[i]>a_quartile[0]+1.5*(a_quartile[2]-a_quartile[0]) or b[i]<b_quartile[0]-1.5*(b_quartile[2]-b_quartile[0]) or b[i]>b_quartile[0]+1.5*(b_quartile[2]-b_quartile[0]):
        ws.cell(row=i+2,column=21).value=None
        ws.cell(row=i+2,column=22).value =None
    elif l[i]<l_quartile[0]-1.5*(l_quartile[2]-l_quartile[0]) or l[i]>l_quartile[0]+1.5*(l_quartile[2]-l_quartile[0]):
        ws.cell(row=i + 2, column=21).value = None
        ws.cell(row=i + 2, column=22).value = None
a=df['L']
b=df['Unnamed: 3']
l=b-a
c=len(a.dropna())
k=0
if c%10<5:
    k=2.752
else:
    k=2.549
a_mean=np.mean(a.dropna())
a_sd=statistics.stdev(a.dropna())
b_mean=np.mean(b.dropna())
b_sd=statistics.stdev(b.dropna())
l_mean=np.mean(l.dropna())
l_sd=statistics.stdev(l.dropna())
for i in range(len(a)):
    if a[i]<a_mean-k*(a_sd) or a[i]>a_mean+k*(a_sd) or b[i]<b_mean-k*(b_sd) or b[i]>b_mean+k*(b_sd):
        ws.cell(row=i+2,column=3).value=None
        ws.cell(row=i+2,column=4).value =None
    elif l[i]<l_mean-k*(l_sd) or l[i]>l_mean+k*(l_sd):
        ws.cell(row=i + 2, column=3).value = None
        ws.cell(row=i + 2, column=4).value = None
a = df['VL']
b = df['Unnamed: 1']
l = b - a
c = len(a.dropna())
k = 0
if c % 10 < 5:
    k = 2.752
else:
    k = 2.549
a_mean = np.mean(a.dropna())
a_sd = statistics.stdev(a.dropna())
b_mean = np.mean(b.dropna())
b_sd = statistics.stdev(b.dropna())
l_mean = np.mean(l.dropna())
l_sd = statistics.stdev(l.dropna())
for i in range(len(a)):
    if a[i]<a_mean-k*(a_sd) or a[i]>a_mean+k*(a_sd) or b[i]<b_mean-k*(b_sd) or b[i]>b_mean+k*(b_sd):
        ws.cell(row=i + 2, column=1).value = None
        ws.cell(row=i + 2, column=2).value = None
    elif l[i]<l_mean -k*(l_sd) or l[i]>l_mean+(k*l_sd):
        ws.cell(row=i + 2, column=1).value = None
        ws.cell(row=i + 2, column=2).value = None
a = df['M']
b = df['Unnamed: 5']
l = b - a
c = len(a.dropna())
k = 0
if c % 10 < 5:
    k = 2.752
else:
    k = 2.549
a_mean = np.mean(a.dropna())
a_sd = statistics.stdev(a.dropna())
b_mean = np.mean(b.dropna())
b_sd = statistics.stdev(b.dropna())
l_mean = np.mean(l.dropna())
l_sd = statistics.stdev(l.dropna())
for i in range(len(a)):
    if a[i]<a_mean-k*(a_sd) or a[i]>a_mean+k*(a_sd) or b[i]<b_mean-k*(b_sd) or b[i]>b_mean+k*(b_sd):
        ws.cell(row=i + 2, column=5).value = None
        ws.cell(row=i + 2, column=6).value = None
    elif l[i]<l_mean -k*(l_sd) or l[i]>l_mean+(k*l_sd):
        ws.cell(row=i + 2, column=5).value = None
        ws.cell(row=i + 2, column=6).value = None
a = df['LH']
b = df['Unnamed: 7']
l = b - a
c = len(a.dropna())
k = 0
if c % 10 < 5:
    k = 2.752
else:
    k = 2.549
a_mean = np.mean(a.dropna())
a_sd = statistics.stdev(a.dropna())
b_mean = np.mean(b.dropna())
b_sd = statistics.stdev(b.dropna())
l_mean = np.mean(l.dropna())
l_sd = statistics.stdev(l.dropna())
for i in range(len(a)):
    if a[i]<a_mean-k*(a_sd) or a[i]>a_mean+k*(a_sd) or b[i]<b_mean-k*(b_sd) or b[i]>b_mean+k*(b_sd):
        ws.cell(row=i + 2, column=7).value = None
        ws.cell(row=i + 2, column=8).value = None
    elif l[i]<l_mean -k*(l_sd) or l[i]>l_mean+(k*l_sd):
        ws.cell(row=i + 2, column=7).value = None
        ws.cell(row=i + 2, column=8).value = None
a = df['H']
b = df['Unnamed: 9']
l = b - a
c = len(a.dropna())
k = 0
if c % 10 < 5:
    k = 2.752
else:
    k = 2.549
a_mean = np.mean(a.dropna())
a_sd = statistics.stdev(a.dropna())
b_mean = np.mean(b.dropna())
b_sd = statistics.stdev(b.dropna())
l_mean = np.mean(l.dropna())
l_sd = statistics.stdev(l.dropna())
for i in range(len(a)):
    if a[i]<a_mean-k*(a_sd) or a[i]>a_mean+k*(a_sd) or b[i]<b_mean-k*(b_sd) or b[i]>b_mean+k*(b_sd):
        ws.cell(row=i + 2, column=9).value = None
        ws.cell(row=i + 2, column=10).value = None
    elif l[i]<l_mean -k*(l_sd) or l[i]>l_mean+(k*l_sd):
        ws.cell(row=i + 2, column=9).value = None
        ws.cell(row=i + 2, column=10).value = None
a = df['VH']
b = df['Unnamed: 11']
l = b - a
c = len(a.dropna())
k = 0
if c % 10 < 5:
    k = 2.752
else:
    k = 2.549
a_mean = np.mean(a.dropna())
a_sd = statistics.stdev(a.dropna())
b_mean = np.mean(b.dropna())
b_sd = statistics.stdev(b.dropna())
l_mean = np.mean(l.dropna())
l_sd = statistics.stdev(l.dropna())
for i in range(len(a)):
    if a[i]<a_mean-k*(a_sd) or a[i]>a_mean+k*(a_sd) or b[i]<b_mean-k*(b_sd) or b[i]>b_mean+k*(b_sd):
        ws.cell(row=i + 2, column=11).value = None
        ws.cell(row=i + 2, column=12).value = None
    elif l[i]<l_mean -k*(l_sd) or l[i]>l_mean+(k*l_sd):
        ws.cell(row=i + 2, column=11).value = None
        ws.cell(row=i + 2, column=12).value = None
a = df['P']
b = df['Unnamed: 13']
l = b - a
c = len(a.dropna())
k = 0
if c % 10 < 5:
    k = 2.752
else:
    k = 2.549
a_mean = np.mean(a.dropna())
a_sd = statistics.stdev(a.dropna())
b_mean = np.mean(b.dropna())
b_sd = statistics.stdev(b.dropna())
l_mean = np.mean(l.dropna())
l_sd = statistics.stdev(l.dropna())
for i in range(len(a)):
    if a[i]<a_mean-k*(a_sd) or a[i]>a_mean+k*(a_sd) or b[i]<b_mean-k*(b_sd) or b[i]>b_mean+k*(b_sd):
        ws.cell(row=i + 2, column=13).value = None
        ws.cell(row=i + 2, column=14).value = None
    elif l[i]<l_mean -k*(l_sd) or l[i]>l_mean+(k*l_sd):
        ws.cell(row=i + 2, column=13).value = None
        ws.cell(row=i + 2, column=14).value = None
a = df['MLP']
b = df['Unnamed: 15']
l = b - a
c = len(a.dropna())
k = 0
if c % 10 < 5:
    k = 2.752
else:
    k = 2.549
a_mean = np.mean(a.dropna())
a_sd = statistics.stdev(a.dropna())
b_mean = np.mean(b.dropna())
b_sd = statistics.stdev(b.dropna())
l_mean = np.mean(l.dropna())
l_sd = statistics.stdev(l.dropna())
for i in range(len(a)):
    if a[i]<a_mean-k*(a_sd) or a[i]>a_mean+k*(a_sd) or b[i]<b_mean-k*(b_sd) or b[i]>b_mean+k*(b_sd):
        ws.cell(row=i + 2, column=15).value = None
        ws.cell(row=i + 2, column=16).value = None
    elif l[i]<l_mean -k*(l_sd) or l[i]>l_mean+(k*l_sd):
        ws.cell(row=i + 2, column=17).value = None
        ws.cell(row=i + 2, column=18).value = None
a = df['ME']
b = df['Unnamed: 17']
l = b - a
c = len(a.dropna())
k = 0
if c % 10 < 5:
    k = 2.752
else:
    k = 2.549
a_mean = np.mean(a.dropna())
a_sd = statistics.stdev(a.dropna())
b_mean = np.mean(b.dropna())
b_sd = statistics.stdev(b.dropna())
l_mean = np.mean(l.dropna())
l_sd = statistics.stdev(l.dropna())
for i in range(len(a)):
    if a[i]<a_mean-k*(a_sd) or a[i]>a_mean+k*(a_sd) or b[i]<b_mean-k*(b_sd) or b[i]>b_mean+k*(b_sd):
        ws.cell(row=i + 2, column=17).value = None
        ws.cell(row=i + 2, column=18).value = None
    elif l[i]<l_mean -k*(l_sd) or l[i]>l_mean+(k*l_sd):
        ws.cell(row=i + 2, column=17).value = None
        ws.cell(row=i + 2, column=18).value = None
a = df['G']
b = df['Unnamed: 19']
l = b - a
c = len(a.dropna())
k = 0
if c % 10 < 5:
    k = 2.752
else:
    k = 2.549
a_mean = np.mean(a.dropna())
a_sd = statistics.stdev(a.dropna())
b_mean = np.mean(b.dropna())
b_sd = statistics.stdev(b.dropna())
l_mean = np.mean(l.dropna())
l_sd = statistics.stdev(l.dropna())
for i in range(len(a)):
    if a[i]<a_mean-k*(a_sd) or a[i]>a_mean+k*(a_sd) or b[i]<b_mean-k*(b_sd) or b[i]>b_mean+k*(b_sd):
        ws.cell(row=i + 2, column=19).value = None
        ws.cell(row=i + 2, column=20).value = None
    elif l[i]<l_mean -k*(l_sd) or l[i]>l_mean+(k*l_sd):
        ws.cell(row=i + 2, column=19).value = None
        ws.cell(row=i + 2, column=20).value = None
a = df['VG']
b = df['Unnamed: 21']
l = b - a
c = len(a.dropna())
k = 0
if c % 10 < 5:
    k = 2.752
else:
    k = 2.549
a_mean = np.mean(a.dropna())
a_sd = statistics.stdev(a.dropna())
b_mean = np.mean(b.dropna())
b_sd = statistics.stdev(b.dropna())
l_mean = np.mean(l.dropna())
l_sd = statistics.stdev(l.dropna())
for i in range(len(a)):
    if a[i]<a_mean-k*(a_sd) or a[i]>a_mean+k*(a_sd) or b[i]<b_mean-k*(b_sd) or b[i]>b_mean+k*(b_sd):
        ws.cell(row=i + 2, column=21).value = None
        ws.cell(row=i + 2, column=22).value = None
    elif l[i]<l_mean -k*(l_sd) or l[i]>l_mean+(k*l_sd):
        ws.cell(row=i + 2, column=21).value = None
        ws.cell(row=i + 2, column=22).value = None
a=df['VL']
b=df['Unnamed: 1']
a_mean = np.mean(a.dropna())
a_sd = (statistics.stdev(a.dropna(),xbar=a_mean))
b_mean = np.mean(b.dropna())
b_sd = (statistics.stdev(b.dropna(),xbar=b_mean))
print(b_sd)
c=float(a_sd)/float(b_sd)
sigma1=(b_mean*a_sd*a_sd-a_mean*b_sd*b_sd)+(a_sd*b_sd)*((a_mean-b_mean)*(a_mean-b_mean)+2*(a_sd**2-b_sd**2)*ln(c))**0.5/float(a_sd**2-b_sd**2)
sigma2=(b_mean*a_sd*a_sd-a_mean*b_sd*b_sd)-(a_sd*b_sd)*((a_mean-b_mean)*(a_mean-b_mean)+2*(a_sd**2-b_sd**2)*ln(c))**0.5/float(a_sd**2-b_sd**2)
for i in range(len(a)):
    if a[i]>sigma1 and a[i]>sigma2 and b[i]<sigma1 and b[i]<sigma2:
        ws.cell(row=i+2,column=1).value=None
        ws.cell(row=i+2,column=2).value =None
a=df['L']
b=df['Unnamed: 3']
a_mean = np.mean(a.dropna())
a_sd = (statistics.stdev(a.dropna(),xbar=a_mean))
b_mean = np.mean(b.dropna())
b_sd = (statistics.stdev(b.dropna(),xbar=b_mean))
c=float(a_sd)/float(b_sd)
sigma1=(b_mean*a_sd*a_sd-a_mean*b_sd*b_sd)+(a_sd*b_sd)*((a_mean-b_mean)*(a_mean-b_mean)+2*(a_sd**2-b_sd**2)*ln(c))**0.5/float(a_sd**2-b_sd**2)
sigma2=(b_mean*a_sd*a_sd-a_mean*b_sd*b_sd)-(a_sd*b_sd)*((a_mean-b_mean)*(a_mean-b_mean)+2*(a_sd**2-b_sd**2)*ln(c))**0.5/float(a_sd**2-b_sd**2)
for i in range(len(a)):
    if a[i]>sigma1 and a[i]>sigma2 and b[i]<sigma1 and b[i]<sigma2:
        ws.cell(row=i+2,column=3).value=None
        ws.cell(row=i+2,column=4).value =None
a=df['M']
b=df['Unnamed: 5']
a_mean = np.mean(a.dropna())
a_sd = (statistics.stdev(a.dropna(),xbar=a_mean))
b_mean = np.mean(b.dropna())
b_sd = (statistics.stdev(b.dropna(),xbar=b_mean))
c=float(a_sd)/float(b_sd)
sigma1=(b_mean*a_sd*a_sd-a_mean*b_sd*b_sd)+(a_sd*b_sd)*((a_mean-b_mean)*(a_mean-b_mean)+2*(a_sd**2-b_sd**2)*ln(c))**0.5/float(a_sd**2-b_sd**2)
sigma2=(b_mean*a_sd*a_sd-a_mean*b_sd*b_sd)-(a_sd*b_sd)*((a_mean-b_mean)*(a_mean-b_mean)+2*(a_sd**2-b_sd**2)*ln(c))**0.5/float(a_sd**2-b_sd**2)
for i in range(len(a)):
    if a[i]>sigma1 and a[i]>sigma2 and b[i]<sigma1 and b[i]<sigma2:
        ws.cell(row=i+2,column=1).value=None
        ws.cell(row=i+2,column=2).value =None
a=df['LH']
b=df['Unnamed: 7']
a_mean = np.mean(a.dropna())
a_sd = (statistics.stdev(a.dropna(),xbar=a_mean))
b_mean = np.mean(b.dropna())
b_sd = (statistics.stdev(b.dropna(),xbar=b_mean))
c=float(a_sd)/float(b_sd)
sigma1=(b_mean*a_sd*a_sd-a_mean*b_sd*b_sd)+(a_sd*b_sd)*((a_mean-b_mean)*(a_mean-b_mean)+2*(a_sd**2-b_sd**2)*ln(c))**0.5/float(a_sd**2-b_sd**2)
sigma2=(b_mean*a_sd*a_sd-a_mean*b_sd*b_sd)-(a_sd*b_sd)*((a_mean-b_mean)*(a_mean-b_mean)+2*(a_sd**2-b_sd**2)*ln(c))**0.5/float(a_sd**2-b_sd**2)
for i in range(len(a)):
    if a[i]>sigma1 and a[i]>sigma2 and b[i]<sigma1 and b[i]<sigma2:
        ws.cell(row=i+2,column=7).value=None
        ws.cell(row=i+2,column=8).value =None
a=df['H']
b=df['Unnamed: 9']
a_mean = np.mean(a.dropna())
a_sd = (statistics.stdev(a.dropna(),xbar=a_mean))
b_mean = np.mean(b.dropna())
b_sd = (statistics.stdev(b.dropna(),xbar=b_mean))
c=float(a_sd)/float(b_sd)
sigma1=(b_mean*a_sd*a_sd-a_mean*b_sd*b_sd)+(a_sd*b_sd)*((a_mean-b_mean)*(a_mean-b_mean)+2*(a_sd**2-b_sd**2)*ln(c))**0.5/float(a_sd**2-b_sd**2)
sigma2=(b_mean*a_sd*a_sd-a_mean*b_sd*b_sd)-(a_sd*b_sd)*((a_mean-b_mean)*(a_mean-b_mean)+2*(a_sd**2-b_sd**2)*ln(c))**0.5/float(a_sd**2-b_sd**2)
for i in range(len(a)):
    if a[i]>sigma1 and a[i]>sigma2 and b[i]<sigma1 and b[i]<sigma2:
        ws.cell(row=i+2,column=9).value=None
        ws.cell(row=i+2,column=10).value =None
a=df['VH']
b=df['Unnamed: 11']
a_mean = np.mean(a.dropna())
a_sd = (statistics.stdev(a.dropna(),xbar=a_mean))
b_mean = np.mean(b.dropna())
b_sd = (statistics.stdev(b.dropna(),xbar=b_mean))
c=float(a_sd)/float(b_sd)
sigma1=(b_mean*a_sd*a_sd-a_mean*b_sd*b_sd)+(a_sd*b_sd)*((a_mean-b_mean)*(a_mean-b_mean)+2*(a_sd**2-b_sd**2)*ln(c))**0.5/float(a_sd**2-b_sd**2)
sigma2=(b_mean*a_sd*a_sd-a_mean*b_sd*b_sd)-(a_sd*b_sd)*((a_mean-b_mean)*(a_mean-b_mean)+2*(a_sd**2-b_sd**2)*ln(c))**0.5/float(a_sd**2-b_sd**2)
for i in range(len(a)):
    if a[i]>sigma1 and a[i]>sigma2 and b[i]<sigma1 and b[i]<sigma2:
        ws.cell(row=i+2,column=11).value=None
        ws.cell(row=i+2,column=12).value =None
a=df['P']
b=df['Unnamed: 13']
a_mean = np.mean(a.dropna())
a_sd = (statistics.stdev(a.dropna(),xbar=a_mean))
b_mean = np.mean(b.dropna())
b_sd = (statistics.stdev(b.dropna(),xbar=b_mean))
c=float(a_sd)/float(b_sd)
sigma1=(b_mean*a_sd*a_sd-a_mean*b_sd*b_sd)+(a_sd*b_sd)*((a_mean-b_mean)*(a_mean-b_mean)+2*(a_sd**2-b_sd**2)*ln(c))**0.5/float(a_sd**2-b_sd**2)
sigma2=(b_mean*a_sd*a_sd-a_mean*b_sd*b_sd)-(a_sd*b_sd)*((a_mean-b_mean)*(a_mean-b_mean)+2*(a_sd**2-b_sd**2)*ln(c))**0.5/float(a_sd**2-b_sd**2)
for i in range(len(a)):
    if a[i]>sigma1 and a[i]>sigma2 and b[i]<sigma1 and b[i]<sigma2:
        ws.cell(row=i+2,column=13).value=None
        ws.cell(row=i+2,column=14).value =None
a=df['MLP']
b=df['Unnamed: 15']
a_mean = np.mean(a.dropna())
a_sd = (statistics.stdev(a.dropna(),xbar=a_mean))
b_mean = np.mean(b.dropna())
b_sd = (statistics.stdev(b.dropna(),xbar=b_mean))
c=float(a_sd)/float(b_sd)
sigma1=(b_mean*a_sd*a_sd-a_mean*b_sd*b_sd)+(a_sd*b_sd)*((a_mean-b_mean)*(a_mean-b_mean)+2*(a_sd**2-b_sd**2)*ln(c))**0.5/float(a_sd**2-b_sd**2)
sigma2=(b_mean*a_sd*a_sd-a_mean*b_sd*b_sd)-(a_sd*b_sd)*((a_mean-b_mean)*(a_mean-b_mean)+2*(a_sd**2-b_sd**2)*ln(c))**0.5/float(a_sd**2-b_sd**2)
for i in range(len(a)):
    if a[i]>sigma1 and a[i]>sigma2 and b[i]<sigma1 and b[i]<sigma2:
        ws.cell(row=i+2,column=15).value=None
        ws.cell(row=i+2,column=16).value =None
a=df['ME']
b=df['Unnamed: 17']
a_mean = np.mean(a.dropna())
a_sd = (statistics.stdev(a.dropna(),xbar=a_mean))
b_mean = np.mean(b.dropna())
b_sd = (statistics.stdev(b.dropna(),xbar=b_mean))
c=float(a_sd)/float(b_sd)
sigma1=(b_mean*a_sd*a_sd-a_mean*b_sd*b_sd)+(a_sd*b_sd)*((a_mean-b_mean)*(a_mean-b_mean)+2*(a_sd**2-b_sd**2)*ln(c))**0.5/float(a_sd**2-b_sd**2)
sigma2=(b_mean*a_sd*a_sd-a_mean*b_sd*b_sd)-(a_sd*b_sd)*((a_mean-b_mean)*(a_mean-b_mean)+2*(a_sd**2-b_sd**2)*ln(c))**0.5/float(a_sd**2-b_sd**2)
for i in range(len(a)):
    if a[i]>sigma1 and a[i]>sigma2 and b[i]<sigma1 and b[i]<sigma2:
        ws.cell(row=i+2,column=17).value=None
        ws.cell(row=i+2,column=18).value =None
a=df['G']
b=df['Unnamed: 19']
a_mean = np.mean(a.dropna())
a_sd = (statistics.stdev(a.dropna(),xbar=a_mean))
b_mean = np.mean(b.dropna())
b_sd = (statistics.stdev(b.dropna(),xbar=b_mean))
c=float(a_sd)/float(b_sd)
sigma1=(b_mean*a_sd*a_sd-a_mean*b_sd*b_sd)+(a_sd*b_sd)*((a_mean-b_mean)*(a_mean-b_mean)+2*(a_sd**2-b_sd**2)*ln(c))**0.5/float(a_sd**2-b_sd**2)
sigma2=(b_mean*a_sd*a_sd-a_mean*b_sd*b_sd)-(a_sd*b_sd)*((a_mean-b_mean)*(a_mean-b_mean)+2*(a_sd**2-b_sd**2)*ln(c))**0.5/float(a_sd**2-b_sd**2)
for i in range(len(a)):
    if a[i]>sigma1 and a[i]>sigma2 and b[i]<sigma1 and b[i]<sigma2:
        ws.cell(row=i+2,column=19).value=None
        ws.cell(row=i+2,column=20).value =None
a=df['VG']
b=df['Unnamed: 21']
a_mean = np.mean(a.dropna())
a_sd = np.std(a.dropna())
b_mean = np.mean(b.dropna())
b_sd = np.std(b.dropna())
c=float(a_sd)/float(b_sd)
sigma1=(b_mean*a_sd*a_sd-a_mean*b_sd*b_sd)+(a_sd*b_sd)*((a_mean-b_mean)*(a_mean-b_mean)+2*(a_sd**2-b_sd**2)*ln(c))**0.5/float(a_sd**2-b_sd**2)
sigma2=(b_mean*a_sd*a_sd-a_mean*b_sd*b_sd)-(a_sd*b_sd)*((a_mean-b_mean)*(a_mean-b_mean)+2*(a_sd**2-b_sd**2)*ln(c))**0.5/float(a_sd**2-b_sd**2)
for i in range(len(a)):
    if a[i]>sigma1 and a[i]>sigma2 and b[i]<sigma1 and b[i]<sigma2:
        ws.cell(row=i+2,column=21).value=None
        ws.cell(row=i+2,column=22).value =None
wb1.save(file_name)





